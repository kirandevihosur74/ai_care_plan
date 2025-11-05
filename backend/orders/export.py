import csv
import io
import logging
from datetime import datetime
from typing import Optional, List
from django.utils import timezone
from django.db.models import Q
from .models import Order, Patient, Provider
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger('orders')

def get_orders_for_export(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    provider_npi: Optional[str] = None,
    diagnosis: Optional[str] = None
) -> List[Order]:
    queryset = Order.objects.select_related('patient', 'provider').all()
    
    if start_date:
        queryset = queryset.filter(created_at__gte=start_date)
    
    if end_date:
        queryset = queryset.filter(created_at__lte=end_date)
    
    if provider_npi:
        queryset = queryset.filter(provider__npi=provider_npi)
    
    if diagnosis:
        queryset = queryset.filter(
            Q(primary_diagnosis=diagnosis) | Q(additional_diagnoses__contains=[diagnosis])
        )
    
    queryset = queryset.order_by('-created_at')
    
    orders = list(queryset)
    logger.info(f"Export query returned {len(orders)} orders")
    if start_date or end_date:
        logger.info(f"Date filter - start: {start_date}, end: {end_date}")
    if provider_npi:
        logger.info(f"Provider NPI filter: {provider_npi}")
    if diagnosis:
        logger.info(f"Diagnosis filter: {diagnosis}")
    
    return orders

def export_to_csv(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    provider_npi: Optional[str] = None,
    diagnosis: Optional[str] = None
) -> str:
    orders = get_orders_for_export(start_date, end_date, provider_npi, diagnosis)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'Order ID',
        'Order Date',
        'Patient MRN',
        'Patient First Name',
        'Patient Last Name',
        'Provider Name',
        'Provider NPI',
        'Primary Diagnosis (ICD-10)',
        'Additional Diagnoses (ICD-10)',
        'Medication Name',
        'Medication History',
        'Care Plan Generated',
        'Care Plan Generated At',
        'Care Plan Length'
    ])
    
    for order in orders:
        additional_diagnoses_str = ', '.join(order.additional_diagnoses) if order.additional_diagnoses else ''
        medication_history_str = ', '.join(order.medication_history) if order.medication_history else ''
        care_plan_generated = 'Yes' if order.care_plan else 'No'
        care_plan_generated_at = order.care_plan_generated_at.strftime('%Y-%m-%d %H:%M:%S') if order.care_plan_generated_at else ''
        care_plan_length = len(order.care_plan) if order.care_plan else 0
        
        writer.writerow([
            order.id,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.patient.mrn,
            order.patient.first_name,
            order.patient.last_name,
            order.provider.name,
            order.provider.npi,
            order.primary_diagnosis,
            additional_diagnoses_str,
            order.medication_name,
            medication_history_str,
            care_plan_generated,
            care_plan_generated_at,
            care_plan_length
        ])
    
    csv_content = output.getvalue()
    output.close()
    
    logger.info(f"CSV export generated with {len(orders)} orders")
    return csv_content

def export_to_excel(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    provider_npi: Optional[str] = None,
    diagnosis: Optional[str] = None
) -> bytes:
    orders = get_orders_for_export(start_date, end_date, provider_npi, diagnosis)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Care Plans Export"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        'Order ID',
        'Order Date',
        'Patient MRN',
        'Patient First Name',
        'Patient Last Name',
        'Provider Name',
        'Provider NPI',
        'Primary Diagnosis (ICD-10)',
        'Additional Diagnoses (ICD-10)',
        'Medication Name',
        'Medication History',
        'Care Plan Generated',
        'Care Plan Generated At',
        'Care Plan Length'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_idx, order in enumerate(orders, 2):
        additional_diagnoses_str = ', '.join(order.additional_diagnoses) if order.additional_diagnoses else ''
        medication_history_str = ', '.join(order.medication_history) if order.medication_history else ''
        care_plan_generated = 'Yes' if order.care_plan else 'No'
        care_plan_generated_at = order.care_plan_generated_at.strftime('%Y-%m-%d %H:%M:%S') if order.care_plan_generated_at else ''
        care_plan_length = len(order.care_plan) if order.care_plan else 0
        
        ws.cell(row=row_idx, column=1, value=order.id)
        ws.cell(row=row_idx, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=3, value=order.patient.mrn)
        ws.cell(row=row_idx, column=4, value=order.patient.first_name)
        ws.cell(row=row_idx, column=5, value=order.patient.last_name)
        ws.cell(row=row_idx, column=6, value=order.provider.name)
        ws.cell(row=row_idx, column=7, value=order.provider.npi)
        ws.cell(row=row_idx, column=8, value=order.primary_diagnosis)
        ws.cell(row=row_idx, column=9, value=additional_diagnoses_str)
        ws.cell(row=row_idx, column=10, value=order.medication_name)
        ws.cell(row=row_idx, column=11, value=medication_history_str)
        ws.cell(row=row_idx, column=12, value=care_plan_generated)
        ws.cell(row=row_idx, column=13, value=care_plan_generated_at)
        ws.cell(row=row_idx, column=14, value=care_plan_length)
    
    summary_row = len(orders) + 3
    care_plans_count = sum(1 for order in orders if order.care_plan)
    
    ws.cell(row=summary_row, column=1, value="Total Orders:").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(orders))
    
    ws.cell(row=summary_row + 1, column=1, value="Care Plans Generated:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=care_plans_count)
    
    for col in range(1, len(headers) + 1):
        column_letter = ws.cell(row=1, column=col).column_letter
        max_length = 0
        for row in ws.iter_rows(min_row=1, max_row=len(orders) + 1, min_col=col, max_col=col):
            cell_value = str(row[0].value) if row[0].value else ''
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()
    output.close()
    
    logger.info(f"Excel export generated with {len(orders)} orders")
    return excel_content

def get_export_filename(
    format: str,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None
) -> str:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if start_date and end_date:
        date_range = f"{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"
        filename = f"care_plans_export_{date_range}_{timestamp}.{format}"
    elif start_date:
        date_range = f"from_{start_date.strftime('%Y%m%d')}"
        filename = f"care_plans_export_{date_range}_{timestamp}.{format}"
    elif end_date:
        date_range = f"until_{end_date.strftime('%Y%m%d')}"
        filename = f"care_plans_export_{date_range}_{timestamp}.{format}"
    else:
        filename = f"care_plans_export_{timestamp}.{format}"
    
    return filename

